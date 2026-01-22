from flask import (
    render_template, request, redirect, url_for,
    flash, jsonify, current_app, send_from_directory, send_file, make_response
)
from flask_login import current_user, login_required
from app.ayudas import ayudas_bp
from app.models import (
    Paciente, HistoriaClinica, AyudaDiagnostica,
    CatLaboratorioExamen, CatLaboratorioParametro,
    LabSolicitud, LabResultado, OrdenMedica, OrdenLaboratorioItem
)
from app.extensions import db
from datetime import datetime
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from app.utils.fechas import ahora_bogota
from werkzeug.utils import secure_filename
from xhtml2pdf import pisa  # Cambiamos pdfkit por pisa
from io import BytesIO

UPLOAD_SUBFOLDER = os.path.join('uploads', 'ayudas')

def guardar_archivo_ayuda(historia_id: int, tipo: str, file_storage):
    """Guarda un archivo para una ayuda diagnóstica y devuelve la ruta RELATIVA."""
    if not file_storage or not file_storage.filename:
        return None

    nombre_seguro = secure_filename(file_storage.filename)
    base_dir = os.path.join(current_app.root_path, UPLOAD_SUBFOLDER, tipo, str(historia_id))
    os.makedirs(base_dir, exist_ok=True)

    ruta_absoluta = os.path.join(base_dir, nombre_seguro)
    file_storage.save(ruta_absoluta)

    ruta_relativa = os.path.relpath(ruta_absoluta, current_app.root_path)
    return ruta_relativa


# ========== RUTAS PRINCIPALES ==========

@ayudas_bp.route('/')
@login_required
def inicio_ayudas():
    """Pantalla principal del módulo: formulario de búsqueda de paciente."""
    return render_template(
        'ayudas/buscar_paciente.html',
        current_user=current_user
    )


@ayudas_bp.route('/buscar', methods=['POST'])
@login_required
def buscar_paciente():
    """Procesa el formulario de búsqueda y muestra resultados."""
    criterio = request.form.get('criterio', '').strip()

    if not criterio:
        flash('Ingrese un criterio de búsqueda (nombre o número de historia/paciente).', 'warning')
        return redirect(url_for('ayudas.inicio_ayudas'))

    pacientes = (
        Paciente.query
        .filter(
            (Paciente.nombre.ilike(f"%{criterio}%")) |
            (Paciente.numero.ilike(f"%{criterio}%"))
        )
        .all()
    )

    if not pacientes:
        flash('No se encontraron pacientes con ese criterio.', 'info')
        return redirect(url_for('ayudas.inicio_ayudas'))

    if len(pacientes) == 1:
        paciente_unico = pacientes[0]
        return redirect(url_for('ayudas.seleccionar_historia', paciente_id=paciente_unico.id))

    return render_template(
        'ayudas/buscar_resultados.html',
        pacientes=pacientes,
        criterio=criterio,
        current_user=current_user
    )


@ayudas_bp.route('/paciente/<int:paciente_id>/historias')
@login_required
def seleccionar_historia(paciente_id):
    """Lista las historias clínicas de un paciente para seleccionar una."""
    paciente = Paciente.query.get_or_404(paciente_id)
    historias = (
        HistoriaClinica.query
        .filter_by(paciente_id=paciente.id)
        .order_by(HistoriaClinica.fecha_registro.desc())
        .all()
    )

    if not historias:
        flash('Este paciente no tiene historias clínicas registradas.', 'info')
        return redirect(url_for('ayudas.inicio_ayudas'))

    return render_template(
        'ayudas/seleccionar_historia.html',
        paciente=paciente,
        historias=historias,
        current_user=current_user
    )


@ayudas_bp.route('/historia/<int:historia_id>/ayudas')
@login_required
def menu_ayudas_historia(historia_id):
    """Menú principal de ayudas diagnósticas para una historia."""
    historia = HistoriaClinica.query.get_or_404(historia_id)
    paciente = historia.paciente

    return render_template(
        'ayudas/menu_ayudas_historia.html',
        paciente=paciente,
        historia=historia,
        current_user=current_user
    )


# ========== SUBMÓDULO: IMÁGENES ==========

@ayudas_bp.route('/historia/<int:historia_id>/imagenes', methods=['GET', 'POST'])
@login_required
def ayudas_imagenes(historia_id):
    historia = HistoriaClinica.query.get_or_404(historia_id)
    paciente = historia.paciente

    if request.method == 'POST':
        if 'nombre_examen' in request.form:
            nombre = request.form.get('nombre_examen', '').strip()
            fecha_str = request.form.get('fecha_resultado', '').strip()
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d') if fecha_str else None
            observaciones = request.form.get('observaciones', '').strip()
            archivo_rel = guardar_archivo_ayuda(historia_id, 'imagenes', request.files.get('archivo'))

            if nombre:
                ayuda = AyudaDiagnostica(
                    historia_id=historia.id,
                    tipo='imagen',
                    nombre_examen=nombre,
                    fecha_resultado=fecha,
                    archivo=archivo_rel,
                    observaciones=observaciones
                )
                db.session.add(ayuda)
                db.session.commit()
                flash('Estudio de imágenes registrado correctamente.', 'success')
            else:
                flash('El nombre del examen es obligatorio.', 'warning')

        elif 'ayuda_id' in request.form:
            ayuda_id = request.form.get('ayuda_id', type=int)
            ayuda = AyudaDiagnostica.query.get_or_404(ayuda_id)
            ayuda.observaciones = request.form.get('observaciones', '').strip()
            db.session.commit()
            flash('Observaciones actualizadas.', 'success')

    examenes = (
        AyudaDiagnostica.query
        .filter_by(historia_id=historia.id, tipo='imagen')
        .order_by(AyudaDiagnostica.fecha_resultado.desc().nullslast())
        .all()
    )

    return render_template(
        'ayudas/imagenes_listar.html',
        paciente=paciente,
        historia=historia,
        examenes=examenes,
        current_user=current_user
    )


# ========== SUBMÓDULO: LABORATORIOS ==========

@ayudas_bp.route('/historia/<int:historia_id>/laboratorios', methods=['GET'])
@login_required
def ayudas_laboratorios(historia_id):
    """Lista las solicitudes de laboratorio de una historia clínica."""
    historia = HistoriaClinica.query.get_or_404(historia_id)
    paciente = historia.paciente

    solicitudes = (
        LabSolicitud.query
        .filter_by(historia_id=historia.id)
        .order_by(LabSolicitud.fecha_solicitud.desc())
        .all()
    )

    return render_template(
        'laboratorio/solicitudes_listar.html',
        paciente=paciente,
        historia=historia,
        solicitudes=solicitudes,
        current_user=current_user
    )


@ayudas_bp.route('/historia/<int:historia_id>/laboratorios/nueva', methods=['GET', 'POST'])
@login_required
def nueva_solicitud_laboratorio(historia_id):
    """Crear una nueva solicitud de laboratorio usando el catálogo."""
    historia = HistoriaClinica.query.get_or_404(historia_id)
    paciente = historia.paciente

    examenes = CatLaboratorioExamen.query.order_by(
        CatLaboratorioExamen.grupo, CatLaboratorioExamen.nombre
    ).all()

    if request.method == 'POST':
        examen_id = request.form.get('examen_id', type=int)
        if not examen_id:
            flash('Debe seleccionar un examen de laboratorio.', 'warning')
            return redirect(url_for('ayudas.nueva_solicitud_laboratorio', historia_id=historia_id))

        examen = CatLaboratorioExamen.query.get_or_404(examen_id)

        solicitud = LabSolicitud(
            historia_id=historia.id,
            fecha_solicitud=ahora_bogota(),
            estado='pendiente'
        )
        db.session.add(solicitud)
        db.session.flush()

        for param in examen.parametros:
            resultado = LabResultado(
                solicitud_id=solicitud.id,
                examen_id=examen.id,
                parametro_id=param.id,
                unidad=param.unidad
            )
            db.session.add(resultado)

        db.session.commit()
        flash('Solicitud de laboratorio creada.', 'success')
        return redirect(url_for('ayudas.ver_solicitud_laboratorio', solicitud_id=solicitud.id))

    return render_template(
        'laboratorio/nueva_solicitud.html',
        paciente=paciente,
        historia=historia,
        examenes=examenes,
        current_user=current_user
    )


@ayudas_bp.route('/laboratorio/solicitud/<int:solicitud_id>', methods=['GET', 'POST'])
@login_required
def ver_solicitud_laboratorio(solicitud_id):
    """Vista maestro-detalle para capturar y ver resultados de una solicitud."""
    solicitud = LabSolicitud.query.get_or_404(solicitud_id)
    historia = solicitud.historia
    paciente = historia.paciente

    if request.method == 'POST':
        for res in solicitud.resultados:
            campo_valor = f"valor_{res.id}"
            campo_interp = f"interp_{res.id}"

            if campo_valor in request.form:
                nuevo_valor = request.form.get(campo_valor, '').strip()
                res.valor = nuevo_valor

                try:
                    v = float(nuevo_valor.replace(',', '.'))
                    p = res.parametro
                    if p.valor_ref_min is not None and p.valor_ref_max is not None:
                        res.flag_fuera_rango = not (p.valor_ref_min <= v <= p.valor_ref_max)
                except ValueError:
                    res.flag_fuera_rango = False

            if campo_interp in request.form:
                res.interpretacion = request.form.get(campo_interp, '').strip()

        solicitud.fecha_resultado = ahora_bogota()
        solicitud.estado = 'interpretado'
        db.session.commit()
        flash('Resultados de laboratorio actualizados.', 'success')
        return redirect(url_for('ayudas.ver_solicitud_laboratorio', solicitud_id=solicitud.id))

    solicitudes_historia = (
        LabSolicitud.query
        .filter_by(historia_id=historia.id)
        .order_by(LabSolicitud.fecha_solicitud.desc())
        .all()
    )

    return render_template(
        'laboratorio/solicitud_detalle.html',
        solicitud=solicitud,
        solicitudes_historia=solicitudes_historia,
        historia=historia,
        paciente=paciente,
        current_user=current_user
    )


# ========== CARGA MASIVA LABORATORIOS ==========

@ayudas_bp.route('/laboratorios/descargar-plantilla', methods=['GET'])
@login_required
def descargar_plantilla_laboratorios():
    """Descarga una plantilla Excel para carga masiva de exámenes."""
    wb = openpyxl.Workbook()

    # HOJA PRINCIPAL: EXÁMENES
    ws = wb.active
    ws.title = "Examenes"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'NUMERO_PACIENTE',
        'EXAMEN',
        'PARAMETRO',
        'VALOR',
        'FECHA_RESULTADO',
        'LABORATORIO'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    pacientes = Paciente.query.order_by(Paciente.numero).all()
    row_num = 2
    for p in pacientes:
        cell = ws.cell(row=row_num, column=1)
        cell.value = p.numero
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        for col_num in range(2, len(headers) + 1):
            c = ws.cell(row=row_num, column=col_num)
            c.border = thin_border
            c.alignment = Alignment(horizontal="left", vertical="center")

        row_num += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    # HOJA DE INSTRUCCIONES
    ws_instruc = wb.create_sheet("Instrucciones", 0)
    ws_instruc.column_dimensions['A'].width = 90

    instruc = [
        "INSTRUCCIONES DE CARGA MASIVA DE EXÁMENES DE LABORATORIO",
        "",
        "1. Hoja \"Examenes\":",
        "   - La columna NUMERO_PACIENTE ya viene prellenada con los pacientes existentes.",
        "   - NO modifique estos números. Si no va a registrar exámenes para un paciente,",
        "     simplemente deje su fila en blanco.",
        "",
        "2. COLUMNAS OBLIGATORIAS POR CADA FILA CON DATOS:",
        "   - NUMERO_PACIENTE: Ya viene prellenado.",
        "   - EXAMEN: Nombre exacto del examen en el sistema.",
        "   - PARAMETRO: Nombre exacto del parámetro del examen.",
        "   - VALOR: Valor del resultado.",
        "   - FECHA_RESULTADO: Fecha en formato DD/MM/YYYY.",
        "   - LABORATORIO: Nombre del laboratorio que realizó el examen.",
        "",
        "3. REGLAS:",
        "   - Puede dejar filas sin exámenes si no va a cargar nada para ese paciente.",
        "   - No cambie ni elimine los números de pacientes prellenados.",
        "   - No agregue filas intermedias vacías entre filas que sí tengan datos.",
        "",
        "4. HOJA \"Catalogos\":",
        "   - Contiene la lista de exámenes y parámetros válidos.",
        "   - Copie y pegue desde allí los valores de EXAMEN y PARAMETRO para evitar errores",
        "     de escritura.",
        "",
        "5. DESPUÉS DE LLENAR:",
        "   - Guardar el archivo como Excel (.xlsx).",
        "   - Ir a Módulo Ayudas > Laboratorios > Carga masiva.",
        "   - Seleccionar este archivo.",
        "   - Hacer clic en \"Procesar carga masiva\".",
    ]

    for idx, linea in enumerate(instruc, 1):
        cell = ws_instruc.cell(row=idx, column=1)
        cell.value = linea
        if idx == 1:
            cell.font = Font(bold=True, size=12)

    # HOJA DE CATÁLOGOS
    ws_cat = wb.create_sheet("Catalogos", 2)

    ws_cat['A1'] = "EXAMEN_ID"
    ws_cat['B1'] = "EXAMEN_NOMBRE"
    ws_cat['C1'] = "GRUPO"
    ws_cat['D1'] = "PARAMETRO_NOMBRE"
    ws_cat['E1'] = "UNIDAD"
    ws_cat['F1'] = "VALOR_REF_MIN"
    ws_cat['G1'] = "VALOR_REF_MAX"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_cat.column_dimensions[col].width = 22

    header_cat_font = Font(bold=True)
    for col in range(1, 8):
        ws_cat.cell(row=1, column=col).font = header_cat_font

    examenes = CatLaboratorioExamen.query.order_by(
        CatLaboratorioExamen.grupo, CatLaboratorioExamen.nombre
    ).all()

    row = 2
    for ex in examenes:
        if ex.parametros:
            for p in ex.parametros:
                ws_cat.cell(row=row, column=1, value=ex.id)
                ws_cat.cell(row=row, column=2, value=ex.nombre)
                ws_cat.cell(row=row, column=3, value=ex.grupo or "")
                ws_cat.cell(row=row, column=4, value=p.nombre)
                ws_cat.cell(row=row, column=5, value=p.unidad or "")
                ws_cat.cell(row=row, column=6, value=p.valor_ref_min)
                ws_cat.cell(row=row, column=7, value=p.valor_ref_max)
                row += 1
        else:
            ws_cat.cell(row=row, column=1, value=ex.id)
            ws_cat.cell(row=row, column=2, value=ex.nombre)
            ws_cat.cell(row=row, column=3, value=ex.grupo or "")
            row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Plantilla_Carga_Masiva_Laboratorios.xlsx'
    )


@ayudas_bp.route('/laboratorios/carga_masiva', methods=['GET', 'POST'])
@login_required
def carga_masiva_laboratorios():
    if request.method == 'POST':
        file = request.files.get('archivo_masivo')
        if not file or file.filename == '':
            flash('Debe seleccionar un archivo.', 'warning')
            return redirect(url_for('ayudas.carga_masiva_laboratorios'))

        try:
            if file.filename.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file, sheet_name='Examenes')
            else:
                df = pd.read_csv(file)

            columnas_req = [
                'NUMERO_PACIENTE', 'EXAMEN', 'PARAMETRO',
                'VALOR', 'FECHA_RESULTADO', 'LABORATORIO'
            ]
            faltantes = [c for c in columnas_req if c not in df.columns]
            if faltantes:
                flash(f'Columnas faltantes: {", ".join(faltantes)}', 'danger')
                return redirect(url_for('ayudas.carga_masiva_laboratorios'))

            creados = 0
            errores = []

            for idx, row in df.iterrows():
                try:
                    num_pac = str(row['NUMERO_PACIENTE']).strip()
                    nombre_exam = str(row['EXAMEN']).strip()
                    nombre_param = str(row['PARAMETRO']).strip()
                    valor = str(row['VALOR']).strip()
                    lab_nombre = str(row.get('LABORATORIO') or '').strip()

                    fecha_res = None
                    if pd.notna(row.get('FECHA_RESULTADO')):
                        try:
                            fecha_res = pd.to_datetime(row['FECHA_RESULTADO'], format='%d/%m/%Y').to_pydatetime()
                        except Exception:
                            try:
                                fecha_res = pd.to_datetime(row['FECHA_RESULTADO']).to_pydatetime()
                            except Exception:
                                errores.append(f"Fila {idx+2}: Fecha inválida")
                                continue

                    if not (num_pac and nombre_exam and nombre_param):
                        errores.append(f"Fila {idx+2}: Falta información obligatoria")
                        continue

                    paciente = Paciente.query.filter_by(numero=num_pac).first()
                    if not paciente:
                        errores.append(f"Fila {idx+2}: Paciente {num_pac} no encontrado")
                        continue

                    historia = (
                        HistoriaClinica.query
                        .filter_by(paciente_id=paciente.id)
                        .order_by(HistoriaClinica.fecha_registro.desc())
                        .first()
                    )
                    if not historia:
                        errores.append(f"Fila {idx+2}: Paciente {num_pac} sin historia clínica")
                        continue

                    examen = CatLaboratorioExamen.query.filter_by(nombre=nombre_exam).first()
                    if not examen:
                        errores.append(f"Fila {idx+2}: Examen '{nombre_exam}' no encontrado")
                        continue

                    param = CatLaboratorioParametro.query.filter_by(
                        examen_id=examen.id,
                        nombre=nombre_param
                    ).first()
                    if not param:
                        errores.append(f"Fila {idx+2}: Parámetro '{nombre_param}' no encontrado")
                        continue

                    solicitud = None
                    if fecha_res:
                        solicitud = (
                            LabSolicitud.query
                            .filter_by(
                                historia_id=historia.id,
                                laboratorio_nombre=lab_nombre
                            )
                            .filter(db.func.date(LabSolicitud.fecha_solicitud) == fecha_res.date())
                            .first()
                        )

                    if not solicitud:
                        solicitud = LabSolicitud(
                            historia_id=historia.id,
                            fecha_solicitud=fecha_res or ahora_bogota(),
                            estado='completado',
                            laboratorio_nombre=lab_nombre
                        )
                        db.session.add(solicitud)
                        db.session.flush()

                    resultado = LabResultado(
                        solicitud_id=solicitud.id,
                        examen_id=examen.id,
                        parametro_id=param.id,
                        valor=valor,
                        unidad=param.unidad
                    )
                    db.session.add(resultado)
                    creados += 1

                except Exception as e:
                    errores.append(f"Fila {idx+2}: {str(e)}")
                    continue

            db.session.commit()

            msg = f'Se cargaron {creados} resultados de laboratorio.'
            if errores:
                msg += f' Se encontraron {len(errores)} errores.'
                flash(msg, 'warning')
                for error in errores[:10]:
                    flash(f"  • {error}", 'warning')
            else:
                flash(msg, 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Error procesando archivo: {e}', 'danger')

        return redirect(url_for('ayudas.carga_masiva_laboratorios'))

    return render_template('laboratorio/carga_masiva.html', current_user=current_user)


# ========== UTILIDADES ==========

@ayudas_bp.route('/autocomplete')
@login_required
def autocomplete_pacientes():
    termino = request.args.get('q', '').strip()

    if not termino:
        return jsonify([])

    pacientes = (
        Paciente.query
        .filter(Paciente.nombre.ilike(f"%{termino}%"))
        .order_by(Paciente.nombre.asc())
        .limit(10)
        .all()
    )

    datos = [
        {"id": p.id, "nombre": p.nombre, "numero": p.numero}
        for p in pacientes
    ]
    return jsonify(datos)


@ayudas_bp.route('/archivo/<int:ayuda_id>')
@login_required
def ver_archivo_ayuda(ayuda_id):
    ayuda = AyudaDiagnostica.query.get_or_404(ayuda_id)
    if not ayuda.archivo:
        flash('Este examen no tiene archivo asociado.', 'warning')
        if ayuda.tipo == 'imagen':
            return redirect(url_for('ayudas.ayudas_imagenes', historia_id=ayuda.historia_id))
        return redirect(url_for('ayudas.ayudas_laboratorios', historia_id=ayuda.historia_id))

    ruta_absoluta = os.path.join(current_app.root_path, ayuda.archivo)
    directorio, nombre = os.path.split(ruta_absoluta)

    return send_from_directory(directorio, nombre)


@ayudas_bp.route('/eliminar/<int:ayuda_id>', methods=['POST'])
@login_required
def eliminar_ayuda(ayuda_id):
    ayuda = AyudaDiagnostica.query.get_or_404(ayuda_id)
    historia_id = ayuda.historia_id
    tipo = ayuda.tipo

    if ayuda.archivo:
        ruta_abs = os.path.join(current_app.root_path, ayuda.archivo)
        if os.path.exists(ruta_abs):
            try:
                os.remove(ruta_abs)
            except OSError:
                pass

    db.session.delete(ayuda)
    db.session.commit()
    flash('La ayuda diagnóstica se eliminó correctamente.', 'success')

    if tipo == 'imagen':
        return redirect(url_for('ayudas.ayudas_imagenes', historia_id=historia_id))
    return redirect(url_for('ayudas.ayudas_laboratorios', historia_id=historia_id))


@ayudas_bp.route('/laboratorio/paciente/<int:historia_id>', methods=['GET', 'POST'])
@login_required
def laboratorio_paciente(historia_id):
    historia = HistoriaClinica.query.get_or_404(historia_id)

    ordenes = OrdenMedica.query.filter_by(historia_id=historia_id).all()
    orden_ids = [o.id for o in ordenes]

    if not orden_ids:
        items = []
    else:
        items = (
            db.session.query(OrdenLaboratorioItem)
            .filter(OrdenLaboratorioItem.orden_id.in_(orden_ids))
            .order_by(OrdenLaboratorioItem.id.desc())
            .all()
        )

    solicitud = LabSolicitud.query.filter_by(historia_id=historia_id).first()
    if solicitud is None:
        solicitud = LabSolicitud(historia_id=historia_id)
        db.session.add(solicitud)
        db.session.commit()

    if request.method == 'POST':
        form = request.form
        for key, value in form.items():
            if not key.startswith('resultado['):
                continue

            try:
                parametro_id = int(key[len('resultado['):-1])
            except ValueError:
                continue

            valor = value.strip()

            res = LabResultado.query.filter_by(
                solicitud_id=solicitud.id,
                parametro_id=parametro_id
            ).first()

            if res is None and valor:
                param = CatLaboratorioParametro.query.get(parametro_id)
                if not param:
                    continue
                res = LabResultado(
                    solicitud_id=solicitud.id,
                    examen_id=param.examen_id,
                    parametro_id=parametro_id,
                    valor=valor,
                    unidad=param.unidad,
                )
                db.session.add(res)
            elif res:
                res.valor = valor or None

        db.session.commit()
        flash('Resultados de laboratorio actualizados', 'success')
        return redirect(url_for('ayudas.laboratorio_paciente', historia_id=historia_id))

    items_con_parametros = []
    resultados_existentes = {
        r.parametro_id: r
        for r in LabResultado.query.filter_by(solicitud_id=solicitud.id).all()
    }

    for it in items:
        examen = it.examen
        parametros = (
            CatLaboratorioParametro.query
            .filter_by(examen_id=it.examen_id)
            .order_by(CatLaboratorioParametro.id)
            .all()
        )

        items_con_parametros.append({
            'item': it,
            'examen': examen,
            'parametros': parametros,
            'resultados': resultados_existentes,
        })

    return render_template(
        'ayudas/laboratorio_paciente.html',
        historia=historia,
        items_con_parametros=items_con_parametros,
        solicitud=solicitud,
    )

# CORRECCIÓN DE LA RUTA PDF: Se usa ayudas_bp y los modelos correctos definidos arriba
@ayudas_bp.route('/historia/<int:historia_id>/generar_pdf_laboratorio')
@login_required
def generar_pdf_laboratorio(historia_id):
    # 1. Obtención de datos (Misma lógica que ya tienes)
    historia = HistoriaClinica.query.get_or_404(historia_id)
    paciente = historia.paciente
    
    ordenes = OrdenMedica.query.filter_by(historia_id=historia_id).all()
    orden_ids = [o.id for o in ordenes]
    
    items = db.session.query(OrdenLaboratorioItem).filter(OrdenLaboratorioItem.orden_id.in_(orden_ids)).all() if orden_ids else []
    solicitud = LabSolicitud.query.filter_by(historia_id=historia_id).first()
    resultados_existentes = {r.parametro_id: r for r in LabResultado.query.filter_by(solicitud_id=solicitud.id).all()} if solicitud else {}
    
    items_con_parametros = []
    for it in items:
        examen = it.examen
        parametros = CatLaboratorioParametro.query.filter_by(examen_id=it.examen_id).order_by(CatLaboratorioParametro.id).all()
        items_con_parametros.append({'examen': examen, 'parametros': parametros, 'resultados': resultados_existentes})

    # 2. Renderizar HTML
    html = render_template('ayudas/informe.html', 
                           historia=historia, 
                           paciente=paciente, 
                           items_con_parametros=items_con_parametros)

    # 3. Generar PDF con xhtml2pdf
    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=buffer)

    # 4. Verificar errores y retornar
    if pisa_status.err:
        flash("Error al crear el PDF", "danger")
        return redirect(url_for('ayudas.laboratorio_paciente', historia_id=historia_id))

    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    nombre_archivo = f"Lab_{paciente.nombre.replace(' ', '_')}.pdf"
    response.headers['Content-Disposition'] = f'inline; filename={nombre_archivo}'
    
    return response